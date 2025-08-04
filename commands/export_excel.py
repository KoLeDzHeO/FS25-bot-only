from __future__ import annotations

import io
import os
from datetime import datetime
from typing import List, Tuple

import discord
from discord import app_commands
from asyncpg import Pool
from openpyxl import Workbook
from openpyxl.styles import Alignment

from utils.logger import log_debug


async def fetch_players(pool: Pool) -> List[Tuple[str, float, datetime]]:
    """Fetch all players sorted by total time."""
    try:
        rows = await pool.fetch(
            """
            SELECT player_name AS nickname,
                   total_hours,
                   updated_at AS last_seen
            FROM player_total_time
            ORDER BY total_hours DESC;
            """
        )
    except Exception as e:
        log_debug(f"[DB] export_excel fetch error: {e}")
        raise
    return [
        (r["nickname"], float(r["total_hours"]), r["last_seen"])
        for r in rows
    ]


async def _handle_command(interaction: discord.Interaction) -> None:
    await interaction.response.defer()
    pool: Pool = interaction.client.db_pool
    try:
        players = await fetch_players(pool)
    except Exception:
        await interaction.followup.send("Ошибка при получении данных.", ephemeral=True)
        return

    if not players:
        await interaction.followup.send("Нет данных.")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Players"
        ws.append(["Никнейм", "Общее время (ч)", "Последнее обновление"])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        for nickname, total_time, last_seen in players:
            formatted_last_seen = last_seen.strftime("%d.%m.%Y %H:%M")
            ws.append([nickname, total_time, formatted_last_seen])

        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
            row[1].alignment = Alignment(horizontal="center")
            row[2].alignment = Alignment(horizontal="center")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
    except Exception as e:
        log_debug(f"[CMD] export_excel build error: {e}")
        await interaction.followup.send("Ошибка при формировании файла.", ephemeral=True)
        return

    filename = os.path.join("", "players.xlsx")
    try:
        await interaction.followup.send(
            file=discord.File(buffer, filename=filename)
        )
    except Exception as e:
        log_debug(f"[CMD] export_excel send error: {e}")
        await interaction.followup.send("Ошибка при отправке файла.", ephemeral=True)


async def setup(tree: app_commands.CommandTree) -> None:
    @tree.command(name="экспорт_excel", description="Экспорт данных игроков в Excel")
    async def export_excel_command(interaction: discord.Interaction) -> None:
        await _handle_command(interaction)

    log_debug("[Slash] Команда /экспорт_excel зарегистрирована")
